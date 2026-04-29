# -*- coding: utf-8 -*-
"""
임상연구 데이터 품질관리 자동화 시스템
Clinical Research Data Quality Management System

IRB·DRB 승인 기반 임상연구에서 병록번호 기준으로 수령한 R-ID와
수집 데이터 간의 정합성을 자동으로 검증·관리합니다.
"""
import sys
import re
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from pathlib import Path
from datetime import date, datetime
from typing import Optional, List, Tuple

import pandas as pd
import numpy as np

warnings.simplefilter(action='ignore', category=FutureWarning)
pd.options.mode.chained_assignment = None

# =============================================================================
# 상수
# =============================================================================
COL_HOSP_NUM  = '병록번호'
COL_NAME      = '이름'
COL_GENDER    = '성별'
COL_BIRTH     = '생년월일'
COL_DATE      = '접수일자'
COL_RID       = 'R-ID'
COL_BCODE     = 'bCODE'
COL_ID_NO     = '식별번호'
COL_PERS_NO   = '개인번호'
COL_GEN_ID    = '생성ID'
COL_NOTE      = '비고'

SHEET_MASTER   = '통합데이터'
SHEET_ERROR    = '오기데이터'
SHEET_FAMILY   = '가족데이터'
SHEET_MATCHED  = '매칭성공'
SHEET_UNMATCH  = '미매칭'
SHEET_DUP      = '중복'
SHEET_HISTORY  = '변경이력'

FAMILY_KEYWORDS = frozenset(['보호자','남편','공여자','뇌사자','자녀','배우자','부인','모친','부친','태아','아기'])
EXCLUDE_KEYWORDS = frozenset(['개명'])

YEAR_THRESHOLD = 24
SUPREME_PREFIX = '슈프림 추출 데이터'


# =============================================================================
# 유틸리티
# =============================================================================

def clean_str(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip()
    return s[:-2] if s.endswith('.0') else s


def pad_hosp_num(val) -> str:
    s = clean_str(val)
    return ('0' + s) if (s.isdigit() and len(s) == 7) else s


def resolve_year(yy: int) -> int:
    if yy > 100:
        return yy
    return 2000 + yy if yy <= YEAR_THRESHOLD else 1900 + yy


def is_family_name(name: str) -> Tuple[bool, str]:
    name = str(name).strip()
    if any(ex in name for ex in EXCLUDE_KEYWORDS):
        return False, name
    for kw in FAMILY_KEYWORDS:
        if kw in name:
            return True, name.replace(kw, '').strip()
    return False, name


def build_internal_key(gender: str, birth_year: int, birth_month: int,
                        name: str, recv_date) -> str:
    """
    내부 식별키(생성ID) 생성
    구성: 성별코드 + 출생연도 + 출생월(2자리) + 이름첫글자 + 접수일자(YYYYMMDD)
    IRB 승인 R-ID와 매칭하기 위한 내부 키값으로만 사용됩니다.
    """
    try:
        d = pd.to_datetime(recv_date, errors='coerce')
        if pd.isna(d):
            return ""
        date_str = d.strftime('%Y%m%d')
        first_char = name[0] if name else ""
        return f"{gender}{birth_year}{birth_month:02d}{first_char}{date_str}"
    except Exception:
        return ""


def parse_personal_no(pers_no: str) -> Tuple[Optional[str], Optional[int], Optional[int]]:
    """개인번호(예: M93.05)에서 성별·출생연도·출생월 파싱"""
    s = clean_str(pers_no).upper()
    if not s:
        return None, None, None
    gender = 'M' if s.startswith('M') else ('F' if s.startswith('F') else None)
    m = re.search(r'(\d{1,2})[.\-](\d{1,2})', s)
    if m:
        yy = int(m.group(1))
        mm = int(m.group(2))
        return gender, resolve_year(yy), mm
    return gender, None, None


def _build_pers_no(gender_raw: str, year: float, month: float) -> str:
    """성별·출생연도·출생월 → 개인번호 형식 (예: F93.05)"""
    g = 'M' if str(gender_raw).strip() in ('M', '남', '남성', '남자') else 'F'
    yy = int(year) % 100
    mm = int(month)
    return f"{g}{yy:02d}.{mm:02d}"


def get_file_path(title: str) -> str:
    root = tk.Tk()
    try:
        root.withdraw()
        root.attributes('-topmost', True)
        return filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm"), ("CSV", "*.csv"), ("All", "*.*")]
        )
    finally:
        root.destroy()


def get_file_paths(title: str) -> List[str]:
    root = tk.Tk()
    try:
        root.withdraw()
        root.attributes('-topmost', True)
        paths = filedialog.askopenfilenames(
            title=title,
            filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm"), ("CSV", "*.csv"), ("All", "*.*")]
        )
        return list(paths) if paths else []
    finally:
        root.destroy()


def read_excel_all_sheets(fpath: Path) -> pd.DataFrame:
    xls = pd.read_excel(fpath, sheet_name=None, engine='calamine')
    dfs = [df for df in xls.values() if not df.empty]
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


# =============================================================================
# 진행 상태 창
# =============================================================================

class ProgressWindow:
    def __init__(self, title: str = "처리 중..."):
        self.root = tk.Toplevel()
        self.root.title(title)
        self.root.geometry("420x100")
        self.root.resizable(False, False)
        self.root.attributes('-topmost', True)
        self._bar = ttk.Progressbar(self.root, length=380, mode='determinate', maximum=100)
        self._bar.pack(pady=10)
        self._label = tk.Label(self.root, text="")
        self._label.pack()
        self._last = -1
        self.root.update()

    def update(self, pct: int, msg: str = ""):
        pct = max(0, min(100, pct))
        if pct != self._last:
            self._bar['value'] = pct
            self._last = pct
        self._label.config(text=msg)
        self.root.update()

    def close(self):
        try:
            self.root.destroy()
        except tk.TclError:
            pass


# =============================================================================
# 1단계: 수집 데이터 통합 및 Master DB 생성
# =============================================================================

def run_build_master_db(progress: ProgressWindow = None):
    """
    다수의 수집일지 파일을 병합하여 통합 Master DB를 생성합니다.
    - 중복 제거 및 식별번호 우선순위 적용 (O > X)
    - 가족/보호자 데이터 자동 분류
    - 오기 데이터 자동 탐지
    """
    if progress:
        progress.update(0, "파일 선택 대기 중...")

    file_paths = get_file_paths("수집일지 파일을 선택하세요 (여러 개 선택 가능)")
    if not file_paths:
        return

    if progress:
        progress.update(10, "파일 읽는 중...")

    dfs = []
    for fp in file_paths:
        try:
            dfs.append(read_excel_all_sheets(Path(fp)))
        except Exception as e:
            print(f"읽기 실패 ({Path(fp).name}): {e}")

    if not dfs:
        messagebox.showerror("오류", "읽을 수 있는 파일이 없습니다.")
        return

    df = pd.concat(dfs, ignore_index=True)

    required = [COL_BCODE, COL_ID_NO, COL_HOSP_NUM, COL_NAME, COL_PERS_NO, COL_DATE]
    missing = [c for c in required if c not in df.columns]
    if missing:
        messagebox.showerror("오류", f"필수 컬럼 누락: {missing}")
        return

    if progress:
        progress.update(30, "데이터 정제 중...")

    df[COL_HOSP_NUM] = df[COL_HOSP_NUM].apply(pad_hosp_num)
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors='coerce').dt.date

    active, family, errors = [], [], []

    for _, row in df.iterrows():
        hosp = clean_str(row.get(COL_HOSP_NUM))
        name = clean_str(row.get(COL_NAME))
        pers = clean_str(row.get(COL_PERS_NO))

        if not hosp:
            continue

        is_fam, clean_name = is_family_name(name)
        row = row.copy()
        row[COL_NAME] = clean_name

        if is_fam:
            row[COL_NOTE] = f"{clean_name}의 가족"
            family.append(row.to_dict())
            continue

        gender, by, bm = parse_personal_no(pers)
        if gender and by and bm:
            recv_date = row.get(COL_DATE)
            gen_id = build_internal_key(gender, by, bm, clean_name, recv_date)
            row[COL_GEN_ID] = gen_id
        else:
            row[COL_GEN_ID] = ""

        active.append(row.to_dict())

    if progress:
        progress.update(60, "중복 제거 및 우선순위 적용 중...")

    df_active = pd.DataFrame(active)
    df_family = pd.DataFrame(family)
    df_error  = pd.DataFrame(errors)

    if not df_active.empty:
        def id_priority(x):
            v = clean_str(x).upper()
            return 1 if v.startswith('O') else (2 if v.startswith('X') else 3)

        df_active['_priority'] = df_active[COL_ID_NO].apply(id_priority)
        df_active['_fill'] = df_active[required].apply(
            lambda r: sum(1 for v in r if clean_str(v)), axis=1)
        df_active = df_active.sort_values(
            ['_priority', '_fill', COL_DATE], ascending=[True, False, False])
        df_active = df_active.drop_duplicates(
            subset=[COL_HOSP_NUM], keep='first').drop(columns=['_priority', '_fill'])

    output_path = Path(file_paths[0]).parent / 'Master_DB.xlsx'

    if progress:
        progress.update(85, "저장 중...")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_active.to_excel(writer, sheet_name=SHEET_MASTER, index=False)
        if not df_error.empty:
            df_error.to_excel(writer, sheet_name=SHEET_ERROR, index=False)
        if not df_family.empty:
            df_family.to_excel(writer, sheet_name=SHEET_FAMILY, index=False)

    if progress:
        progress.update(100, "완료!")

    messagebox.showinfo("완료", f"Master DB 생성 완료!\n저장 위치: {output_path.name}\n"
                                f"- 통합: {len(df_active)}건\n"
                                f"- 가족: {len(df_family)}건\n"
                                f"- 오기: {len(df_error)}건")


# =============================================================================
# 2단계: 병록번호·접수일자 추출 (R-ID 요청용)
# =============================================================================

def run_extract_for_rid_request(progress: ProgressWindow = None):
    """
    Master DB에서 병록번호와 접수일자를 추출합니다.
    IRB·DRB 승인 후 외부 시스템(슈프림 등)에 R-ID를 요청하기 위한 파일을 생성합니다.
    """
    if progress:
        progress.update(0, "파일 선택 대기 중...")

    src = get_file_path("Master_DB.xlsx를 선택하세요")
    if not src:
        return

    if progress:
        progress.update(30, "데이터 읽는 중...")

    df = pd.read_excel(src, sheet_name=SHEET_MASTER, engine='calamine')
    df[COL_HOSP_NUM] = df[COL_HOSP_NUM].apply(pad_hosp_num)
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors='coerce').dt.date

    df_out = df[[COL_HOSP_NUM, COL_DATE]].drop_duplicates()

    save_path = Path(src).parent / '병록번호_접수일자_추출.xlsx'

    if progress:
        progress.update(80, "저장 중...")

    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='추출', index=False, header=False)
        from openpyxl.styles import numbers as _n
        ws = writer.sheets['추출']
        for r in range(1, ws.max_row + 1):
            ws.cell(r, 1).number_format = _n.FORMAT_TEXT

    if progress:
        progress.update(100, "완료!")

    messagebox.showinfo("완료", f"추출 완료!\n저장 위치: {save_path.name}\n추출 건수: {len(df_out):,}건")


# =============================================================================
# 3단계: IRB 승인 R-ID 파일과 Master DB 매칭
# =============================================================================

def _detect_columns(df: pd.DataFrame) -> dict:
    """R-ID 파일의 컬럼을 자동으로 감지합니다."""
    col_map = {}
    for col in df.columns:
        c = str(col).replace(' ', '')
        if any(k in c for k in ['환자ID', 'R-ID', 'RID', '연구별']):
            col_map['rid'] = col
        elif any(k in c for k in ['환자명', '이름', '성명']):
            col_map['name'] = col
        elif '성별' in c:
            col_map['gender'] = col
        elif any(k in c for k in ['생년월일', '생년']):
            col_map['birth'] = col
        elif any(k in c for k in ['방문일', '접수일', '검사일', '기준일']):
            col_map['date'] = col
    return col_map


def _build_rid_map(df_raw: pd.DataFrame, col_map: dict) -> Tuple[dict, set]:
    """R-ID 파일에서 내부식별키 → R-ID 매핑을 구축합니다."""
    rid_map = {}
    dup_keys = set()

    for _, row in df_raw.iterrows():
        rid   = clean_str(row.get(col_map['rid']))
        name  = clean_str(row.get(col_map['name']))
        g_raw = clean_str(row.get(col_map['gender'])).upper()
        birth = clean_str(row.get(col_map['birth']))
        d_raw = clean_str(row.get(col_map['date']))

        if not all([rid, name, g_raw, birth, d_raw]):
            continue

        gender = 'M' if g_raw in ('M','남','남성') else ('F' if g_raw in ('F','W','여','여성') else None)
        if not gender:
            continue

        try:
            bp = re.split(r'[-/.]', birth.strip())
            by = int(bp[0])
            bm = int(bp[1]) if len(bp) >= 2 else 1
            if by < 100:
                by = resolve_year(by)
        except (ValueError, IndexError):
            continue

        try:
            vdt = pd.to_datetime(d_raw, errors='coerce')
            if pd.isna(vdt):
                continue
            date_str = vdt.strftime('%Y%m%d')
        except Exception:
            continue

        key = f"{gender}{by}{bm:02d}{name[0]}{date_str}"
        if key in rid_map:
            dup_keys.add(key)
        else:
            rid_map[key] = rid

    for k in dup_keys:
        rid_map.pop(k, None)

    return rid_map, dup_keys


def run_rid_matching(progress: ProgressWindow = None):
    """
    IRB·DRB 승인 하에 수령한 R-ID 파일과 Master DB를 매칭합니다.
    내부 식별키(생성ID) 기반으로 정합성을 검증하고 R-ID를 할당합니다.
    """
    if progress:
        progress.update(0, "파일 선택 대기 중...")

    src = get_file_path("[매칭] Master DB를 선택하세요")
    if not src:
        return

    if progress:
        progress.update(5, "R-ID 파일 선택 대기 중...")

    raw_paths = get_file_paths("[매칭] R-ID 파일을 선택하세요 (여러 파일 선택 가능)")
    if not raw_paths:
        return

    try:
        if progress:
            progress.update(15, "Master DB 읽는 중...")

        df_master = pd.read_excel(src, sheet_name=SHEET_MASTER, engine='calamine')
        df_master[COL_HOSP_NUM] = df_master[COL_HOSP_NUM].apply(pad_hosp_num)
        df_master[COL_GEN_ID] = df_master[COL_GEN_ID].apply(clean_str)
        if COL_RID not in df_master.columns:
            df_master[COL_RID] = ""
        df_master[COL_RID] = df_master[COL_RID].apply(clean_str)

        if progress:
            progress.update(30, "R-ID 파일 읽는 중...")

        raw_dfs = []
        for rp in raw_paths:
            try:
                raw_dfs.append(pd.read_excel(Path(rp), engine='calamine'))
            except Exception as e:
                print(f"읽기 실패 ({Path(rp).name}): {e}")

        if not raw_dfs:
            messagebox.showerror("오류", "읽을 수 있는 R-ID 파일이 없습니다.")
            return

        df_raw = pd.concat(raw_dfs, ignore_index=True)

        if progress:
            progress.update(45, "컬럼 감지 중...")

        col_map = _detect_columns(df_raw)
        missing = [k for k in ['rid','name','gender','birth','date'] if k not in col_map]
        if missing:
            messagebox.showerror("오류", f"필수 컬럼 없음: {missing}\n파일 컬럼: {list(df_raw.columns)}")
            return

        if progress:
            progress.update(55, "내부식별키 매핑 구축 중...")

        rid_map, dup_keys = _build_rid_map(df_raw, col_map)

        if progress:
            progress.update(70, "매칭 중...")

        matched, unmatched, dups = [], [], []

        for _, row in df_master.iterrows():
            if clean_str(row.get(COL_RID)):
                matched.append(row.to_dict())
                continue

            key = clean_str(row.get(COL_GEN_ID))
            rd = row.to_dict()

            if key in dup_keys:
                rd[COL_NOTE] = '내부식별키 중복 (R-ID 파일 내 동일 키 다수 존재)'
                dups.append(rd)
            elif key and key in rid_map:
                rd[COL_RID] = rid_map[key]
                matched.append(rd)
            else:
                unmatched.append(rd)

        if progress:
            progress.update(85, "저장 중...")

        result_path = Path(src).parent / '매칭결과.xlsx'

        def to_df(rows):
            df = pd.DataFrame(rows)
            if df.empty:
                return pd.DataFrame(columns=list(df_master.columns))
            cols = [c for c in df_master.columns if c in df.columns]
            return df[cols]

        with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
            to_df(matched).to_excel(writer, sheet_name=SHEET_MATCHED, index=False)
            to_df(unmatched).drop(columns=[COL_RID], errors='ignore').to_excel(
                writer, sheet_name=SHEET_UNMATCH, index=False)
            to_df(dups).drop(columns=[COL_RID], errors='ignore').to_excel(
                writer, sheet_name=SHEET_DUP, index=False)

        if progress:
            progress.update(100, "완료!")

        messagebox.showinfo("매칭 완료",
            f"완료!\n"
            f"- 매칭 성공: {len(matched):,}건\n"
            f"- 미매칭:    {len(unmatched):,}건\n"
            f"- 중복:      {len(dups):,}건\n"
            f"- 저장 위치: {result_path.name}")

    except Exception as e:
        messagebox.showerror("오류", f"매칭 중 오류: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# 4단계: 오기입 보정 매칭 (외부 데이터소스 기반)
# =============================================================================

def run_correction_matching(progress: ProgressWindow = None):
    """
    외부 데이터소스(Hubis) 기준으로 성별·생년월일 오기입을 보정한 후 재매칭합니다.
    수기 입력 과정의 휴먼에러로 인해 미매칭된 데이터를 복구합니다.
    """
    if progress:
        progress.update(0, "매칭결과 파일 선택 대기 중...")

    result_src = get_file_path("[보정매칭] 매칭결과 파일을 선택하세요")
    if not result_src:
        return

    if progress:
        progress.update(10, "외부 데이터소스 파일 선택 대기 중...")

    ref_src = get_file_path("[보정매칭] 외부 데이터소스(Hubis) 파일을 선택하세요")
    if not ref_src:
        return

    if progress:
        progress.update(20, "R-ID 파일 선택 대기 중...")

    raw_paths = get_file_paths("[보정매칭] R-ID 파일을 선택하세요 (여러 파일 선택 가능)")
    if not raw_paths:
        return

    try:
        if progress:
            progress.update(30, "미매칭 데이터 읽는 중...")

        mxls = pd.ExcelFile(result_src, engine='calamine')
        if SHEET_UNMATCH not in mxls.sheet_names:
            messagebox.showerror("오류", f"'{SHEET_UNMATCH}' 시트가 없습니다.")
            return
        df_unmatch = pd.read_excel(mxls, SHEET_UNMATCH)

        if progress:
            progress.update(40, "외부 데이터소스 읽는 중...")

        df_ref = pd.read_excel(ref_src, engine='calamine')
        df_ref = df_ref.iloc[1:].reset_index(drop=True)
        df_ref.columns = [COL_BCODE, '성별', '출생연도', '출생월']
        df_ref[COL_BCODE] = df_ref[COL_BCODE].apply(clean_str)
        df_ref['출생연도'] = pd.to_numeric(df_ref['출생연도'], errors='coerce')
        df_ref['출생월']  = pd.to_numeric(df_ref['출생월'], errors='coerce')
        ref_map = {row[COL_BCODE]: row for _, row in df_ref.iterrows()}

        if progress:
            progress.update(50, "오기입 보정 중...")

        corrected = 0
        for idx, row in df_unmatch.iterrows():
            bcode = clean_str(row.get(COL_BCODE, ''))
            if not bcode or bcode not in ref_map:
                continue
            ref = ref_map[bcode]
            ref_gender = 'M' if str(ref['성별']).strip() in ('M','남','남성','남자') else 'F'
            ref_year = int(ref['출생연도']) if pd.notna(ref['출생연도']) else None
            ref_month = int(ref['출생월']) if pd.notna(ref['출생월']) else None
            if not (ref_year and ref_month):
                continue
            name = clean_str(row.get(COL_NAME, ''))
            recv_date = row.get(COL_DATE)
            new_key = build_internal_key(ref_gender, ref_year, ref_month, name, recv_date)
            if new_key:
                df_unmatch.at[idx, COL_GEN_ID] = new_key
                corrected += 1

        if progress:
            progress.update(60, "R-ID 파일 읽는 중...")

        raw_dfs = []
        for rp in raw_paths:
            try:
                raw_dfs.append(pd.read_excel(Path(rp), engine='calamine'))
            except Exception as e:
                print(f"읽기 실패 ({Path(rp).name}): {e}")

        if not raw_dfs:
            messagebox.showerror("오류", "읽을 수 있는 R-ID 파일이 없습니다.")
            return

        df_raw = pd.concat(raw_dfs, ignore_index=True)
        col_map = _detect_columns(df_raw)
        missing = [k for k in ['rid','name','gender','birth','date'] if k not in col_map]
        if missing:
            messagebox.showerror("오류", f"필수 컬럼 없음: {missing}")
            return

        rid_map, dup_keys = _build_rid_map(df_raw, col_map)

        if progress:
            progress.update(75, "재매칭 중...")

        re_matched, still_unmatch, dups = [], [], []
        for _, row in df_unmatch.iterrows():
            key = clean_str(row.get(COL_GEN_ID))
            rd = row.to_dict()
            if key in dup_keys:
                rd[COL_NOTE] = '내부식별키 중복'
                dups.append(rd)
            elif key and key in rid_map:
                rd[COL_RID] = rid_map[key]
                re_matched.append(rd)
            else:
                still_unmatch.append(rd)

        if progress:
            progress.update(88, "저장 중...")

        save_path = Path(result_src).parent / '보정매칭결과.xlsx'

        def to_df(rows):
            df = pd.DataFrame(rows)
            return df if not df.empty else pd.DataFrame()

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            to_df(re_matched).to_excel(writer, sheet_name=SHEET_MATCHED, index=False)
            to_df(still_unmatch).drop(columns=[COL_RID], errors='ignore').to_excel(
                writer, sheet_name=SHEET_UNMATCH, index=False)
            to_df(dups).drop(columns=[COL_RID], errors='ignore').to_excel(
                writer, sheet_name=SHEET_DUP, index=False)

        if progress:
            progress.update(100, "완료!")

        messagebox.showinfo("보정매칭 완료",
            f"완료!\n"
            f"- 오기입 보정: {corrected:,}건\n"
            f"- 재매칭 성공: {len(re_matched):,}건\n"
            f"- 미매칭:      {len(still_unmatch):,}건\n"
            f"- 중복:        {len(dups):,}건\n"
            f"- 저장 위치:   {save_path.name}")

    except Exception as e:
        messagebox.showerror("오류", f"보정매칭 중 오류: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# 5단계: 오류 알림 발송 (메일 미리보기)
# =============================================================================

class _EmailPreviewDialog:
    def __init__(self, to_addr: str, subject: str, body: str):
        self.confirmed = False
        self._root = tk.Tk()
        self._root.withdraw()
        top = tk.Toplevel(self._root)
        top.title("발송 미리보기")
        top.geometry("520x420")
        top.resizable(False, False)
        top.attributes('-topmost', True)

        tk.Label(top, text=f"수신: {to_addr}", anchor='w',
                 font=("맑은 고딕", 9)).pack(fill='x', padx=14, pady=(12, 2))
        tk.Label(top, text=f"제목: {subject}", anchor='w',
                 font=("맑은 고딕", 9)).pack(fill='x', padx=14, pady=2)
        ttk.Separator(top, orient='horizontal').pack(fill='x', padx=14, pady=6)

        frm = tk.Frame(top)
        frm.pack(fill='both', expand=True, padx=14)
        sb = tk.Scrollbar(frm)
        sb.pack(side='right', fill='y')
        txt = tk.Text(frm, yscrollcommand=sb.set, font=("맑은 고딕", 9),
                      wrap='word', relief='flat', bg='#f9f9f9')
        txt.insert('1.0', body)
        txt.config(state='disabled')
        txt.pack(fill='both', expand=True)
        sb.config(command=txt.yview)

        bf = tk.Frame(top)
        bf.pack(pady=10)

        def _confirm():
            self.confirmed = True
            top.destroy()

        tk.Button(bf, text="발송 확인", width=12, bg="#D4F0D4",
                  font=("맑은 고딕", 9), command=_confirm).pack(side='left', padx=8)
        tk.Button(bf, text="취소", width=12, font=("맑은 고딕", 9),
                  command=top.destroy).pack(side='left', padx=8)

        top.grab_set()
        top.wait_window()
        try:
            self._root.destroy()
        except tk.TclError:
            pass


def _ask_target_system() -> Optional[str]:
    result = [None]
    root = tk.Tk()
    root.withdraw()
    top = tk.Toplevel(root)
    top.title("발송 대상 선택")
    top.geometry("260x130")
    top.resizable(False, False)
    top.attributes('-topmost', True)

    tk.Label(top, text="오류 알림을 보낼 시스템을 선택하세요",
             font=("맑은 고딕", 9)).pack(pady=(16, 10))
    bf = tk.Frame(top)
    bf.pack()

    def select(val):
        result[0] = val
        top.destroy()

    tk.Button(bf, text="휴비스쌤", width=10, bg="#D4F0F0",
              font=("맑은 고딕", 9), command=lambda: select('휴비스쌤')).pack(side='left', padx=8)
    tk.Button(bf, text="슈프림", width=10, bg="#F0E6D4",
              font=("맑은 고딕", 9), command=lambda: select('슈프림')).pack(side='left', padx=8)

    top.grab_set()
    top.wait_window()
    try:
        root.destroy()
    except tk.TclError:
        pass
    return result[0]


def _build_email_body(system: str, col_label: str, ids: list, filename: str) -> str:
    id_lines = '\n'.join(f"  - {v}" for v in ids[:50])
    suffix = f"\n  ... 외 {len(ids) - 50}건" if len(ids) > 50 else ""
    return (
        f"안녕하세요.\n\n"
        f"아래 {system} {col_label} 항목에서 데이터 불일치가 확인되었습니다.\n\n"
        f"[출처 파일] {filename}\n"
        f"[미매칭 {col_label}] 총 {len(ids)}건\n\n"
        f"{id_lines}{suffix}\n\n"
        f"해당 항목의 성별·생년월일·이름 초성 정보 확인 및 수정을 요청드립니다.\n\n"
        f"감사합니다."
    )


def run_send_error_notification(progress: ProgressWindow = None):
    """
    미매칭 데이터 목록을 담당자에게 이메일로 알립니다.
    - 휴비스쌤 선택: 미매칭 bCODE 목록
    - 슈프림 선택: 미매칭 병록번호 목록
    데모 환경에서는 실제 발송 없이 미리보기만 표시합니다.
    """
    if progress:
        progress.update(0, "매칭결과 파일 선택 대기 중...")

    src = get_file_path("매칭결과 파일을 선택하세요")
    if not src:
        return

    system = _ask_target_system()
    if not system:
        return

    if progress:
        progress.update(25, "미매칭 데이터 읽는 중...")

    try:
        df = pd.read_excel(src, sheet_name=SHEET_UNMATCH, engine='calamine')
    except Exception as e:
        messagebox.showerror("오류", f"파일 읽기 실패: {e}")
        return

    if df.empty:
        messagebox.showinfo("알림", "미매칭 데이터가 없습니다.")
        return

    col = COL_BCODE if system == '휴비스쌤' else COL_HOSP_NUM
    col_label = 'bCODE' if system == '휴비스쌤' else '병록번호'

    if col not in df.columns:
        messagebox.showerror("오류", f"'{col}' 컬럼이 없습니다.")
        return

    ids = df[col].apply(clean_str).replace('', pd.NA).dropna().unique().tolist()
    if not ids:
        messagebox.showinfo("알림", f"발송할 {col_label}가 없습니다.")
        return

    root = tk.Tk()
    root.withdraw()
    to_addr = simpledialog.askstring("수신자 이메일", "이메일 주소를 입력하세요:", parent=root)
    try:
        root.destroy()
    except tk.TclError:
        pass

    if not to_addr or not to_addr.strip():
        return
    to_addr = to_addr.strip()

    if progress:
        progress.update(60, "미리보기 준비 중...")

    subject = f"[데이터 오류 알림] {system} 미매칭 건 확인 요청 ({date.today()})"
    body = _build_email_body(system, col_label, ids, Path(src).name)
    dlg = _EmailPreviewDialog(to_addr, subject, body)

    if dlg.confirmed:
        # 실제 환경: smtp.company.com:587 연결 후 발송
        # 데모 환경: 미리보기 확인으로 대체
        if progress:
            progress.update(100, "완료!")
        messagebox.showinfo("발송 완료",
            f"[데모] 발송이 완료되었습니다.\n"
            f"수신: {to_addr}\n"
            f"대상: {system} {col_label} {len(ids):,}건\n\n"
            f"※ 실제 환경에서는 회사 SMTP 서버를 통해 발송됩니다.")


# =============================================================================
# 6단계: 수집일지 자동 수정 (수정파일 반영)
# =============================================================================

def _read_hubis_fix(path: Path) -> pd.DataFrame:
    """
    휴비스쌤 수정파일 읽기
    형식: bCODE(KBN_DONOR) | 성별(SEX) | 출생연도(BIRTH_YEAR) | 출생월(BIRTH_MONTH)
    첫 번째 데이터 행은 내부 필드명 행이므로 건너뜁니다.
    """
    try:
        df = pd.read_excel(path, engine='calamine')
        df = df.iloc[1:].reset_index(drop=True)
        df.columns = [COL_BCODE, COL_GENDER, '출생연도', '출생월']
        df[COL_BCODE] = df[COL_BCODE].apply(clean_str)
        df['출생연도'] = pd.to_numeric(df['출생연도'], errors='coerce')
        df['출생월'] = pd.to_numeric(df['출생월'], errors='coerce')
        return df
    except Exception:
        return pd.DataFrame()


def _read_supreme_fix(path: Path) -> pd.DataFrame:
    """슈프림 수정파일 읽기 — '슈프림 추출 데이터(xxx)' 형식, R-ID 대신 병록번호"""
    try:
        xls = pd.ExcelFile(path, engine='calamine')
        target = next(
            (s for s in xls.sheet_names if SUPREME_PREFIX in s),
            xls.sheet_names[0]
        )
        df = pd.read_excel(xls, sheet_name=target)

        rename_map = {}
        for col in df.columns:
            c = str(col)
            if any(k in c for k in ['병록', '환자번호']):
                rename_map[col] = COL_HOSP_NUM
            elif '성별' in c:
                rename_map[col] = COL_GENDER
            elif any(k in c for k in ['생년월일', '생년']):
                rename_map[col] = COL_BIRTH
            elif any(k in c for k in ['이름', '환자명', '성명']):
                rename_map[col] = COL_NAME

        return df.rename(columns=rename_map)
    except Exception:
        return pd.DataFrame()


def run_update_collection_log(progress: ProgressWindow = None):
    """
    수정파일을 기반으로 원본 수집일지를 자동으로 업데이트합니다.
    병록번호를 기준으로 매칭하여 변경된 값만 덮어쓰고 원본 포맷으로 출력합니다.
    - 휴비스쌤: 원본과 동일한 포맷의 수정파일
    - 슈프림: '슈프림 추출 데이터(xxx)' 형식 (R-ID 대신 병록번호)
    """
    if progress:
        progress.update(0, "원본 수집일지 파일 선택 대기 중...")

    orig_paths = get_file_paths("원본 수집일지 파일을 선택하세요 (여러 개 선택 가능)")
    if not orig_paths:
        return

    system = _ask_target_system()
    if not system:
        return

    if progress:
        progress.update(10, "수정파일 선택 대기 중...")

    fix_path = get_file_path(f"[{system}] 수정파일을 선택하세요")
    if not fix_path:
        return

    try:
        if progress:
            progress.update(25, "원본 수집일지 읽는 중...")

        orig_data: dict = {}
        for fp in orig_paths:
            p = Path(fp)
            orig_data[p] = pd.read_excel(p, sheet_name=None, engine='calamine')

        if progress:
            progress.update(40, "수정파일 읽는 중...")

        if system == '휴비스쌤':
            df_fix = _read_hubis_fix(Path(fix_path))
            if df_fix.empty or COL_BCODE not in df_fix.columns:
                messagebox.showerror("오류", "수정파일에서 bCODE를 찾을 수 없습니다.")
                return
            fix_map = {row[COL_BCODE]: row for _, row in df_fix.iterrows()}
            match_col = COL_BCODE
        else:
            df_fix = _read_supreme_fix(Path(fix_path))
            if df_fix.empty or COL_HOSP_NUM not in df_fix.columns:
                messagebox.showerror("오류", "수정파일에서 병록번호를 찾을 수 없습니다.")
                return
            df_fix[COL_HOSP_NUM] = df_fix[COL_HOSP_NUM].apply(pad_hosp_num)
            fix_map = {row[COL_HOSP_NUM]: row for _, row in df_fix.iterrows()}
            match_col = COL_HOSP_NUM

        if progress:
            progress.update(60, "수정 사항 적용 중...")

        total_updated = 0
        saved = []

        for orig_path, sheets in orig_data.items():
            updated_sheets = {}

            for sheet_name, df in sheets.items():
                if match_col not in df.columns:
                    updated_sheets[sheet_name] = df
                    continue

                df = df.copy()
                df[match_col] = df[match_col].apply(
                    pad_hosp_num if match_col == COL_HOSP_NUM else clean_str)

                if system == '휴비스쌤':
                    for idx, row in df.iterrows():
                        key = clean_str(row.get(match_col, ''))
                        if not key or key not in fix_map:
                            continue
                        fix_row = fix_map[key]
                        year = fix_row.get('출생연도')
                        month = fix_row.get('출생월')
                        gender = fix_row.get(COL_GENDER)
                        if pd.notna(year) and pd.notna(month) and pd.notna(gender):
                            df.at[idx, COL_PERS_NO] = _build_pers_no(
                                str(gender), float(year), float(month))
                            total_updated += 1
                else:
                    fix_cols = [c for c in df_fix.columns
                                if c in df.columns and c != match_col]
                    for idx, row in df.iterrows():
                        key = clean_str(row.get(match_col, ''))
                        if not key or key not in fix_map:
                            continue
                        fix_row = fix_map[key]
                        for col in fix_cols:
                            new_val = fix_row.get(col)
                            if pd.notna(new_val) and str(new_val).strip():
                                df.at[idx, col] = new_val
                                total_updated += 1

                updated_sheets[sheet_name] = df

            out_path = orig_path.parent / f"{orig_path.stem}_수정완료{orig_path.suffix}"
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                for sheet_name, df in updated_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    if COL_HOSP_NUM in df.columns:
                        from openpyxl.styles import numbers as _n
                        ws = writer.sheets[sheet_name]
                        col_idx = list(df.columns).index(COL_HOSP_NUM) + 1
                        for r in range(2, ws.max_row + 1):
                            ws.cell(r, col_idx).number_format = _n.FORMAT_TEXT

            saved.append(out_path.name)

        if progress:
            progress.update(100, "완료!")

        messagebox.showinfo("완료",
            f"수집일지 자동 수정 완료!\n"
            f"- 처리 파일: {len(saved)}개\n"
            f"- 수정된 셀: {total_updated:,}개\n"
            f"- 저장: {'  ·  '.join(saved)}")

    except Exception as e:
        messagebox.showerror("오류", f"수집일지 수정 중 오류: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# GUI
# =============================================================================

def main():
    root = tk.Tk()
    root.title("임상연구 데이터 품질관리 시스템")
    root.geometry("380x360")
    root.resizable(False, False)

    tk.Label(root, text="임상연구 데이터 품질관리 시스템",
             font=("맑은 고딕", 13, "bold")).pack(pady=(18, 4))
    tk.Label(root, text="IRB·DRB 승인 기반 R-ID 매칭 및 데이터 정합성 검증",
             font=("맑은 고딕", 9), fg="#555").pack(pady=(0, 14))

    frm = tk.Frame(root)
    frm.pack()

    btn_opts = [
        ("1. Master DB 생성",            "#D4F0F0", run_build_master_db,           "Master DB 생성 중..."),
        ("2. 병록번호·접수일자 추출",    "#D4F0D4", run_extract_for_rid_request,   "병록번호·접수일자 추출 중..."),
        ("3. R-ID 매칭",                 "#F0E6D4", run_rid_matching,              "R-ID 매칭 중..."),
        ("4. 오기입 보정 매칭",          "#F0D4E6", run_correction_matching,       "오기입 보정 매칭 중..."),
        ("5. 오류 알림 발송",            "#FFFBD4", run_send_error_notification,   "오류 알림 준비 중..."),
        ("6. 수집일지 자동 수정",        "#EAD4F0", run_update_collection_log,     "수집일지 자동 수정 중..."),
    ]

    def run_task(func, title):
        progress_root = tk.Tk()
        progress_root.withdraw()
        progress = ProgressWindow(title)
        try:
            func(progress=progress)
        except Exception as e:
            messagebox.showerror("오류", str(e))
        finally:
            progress.close()
            try:
                progress_root.destroy()
            except tk.TclError:
                pass

    for txt, col, func, title in btn_opts:
        tk.Button(
            frm, text=txt, width=32, height=1, bg=col,
            font=("맑은 고딕", 10),
            command=lambda f=func, t=title: run_task(f, t)
        ).pack(pady=4)

    root.mainloop()


if __name__ == "__main__":
    main()
